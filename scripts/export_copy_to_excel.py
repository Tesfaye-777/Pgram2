import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "游戏文案整理.xlsx"


def read(rel: str) -> str:
    return (ROOT / rel).read_text(encoding="utf-8")


def qval(block: str, key: str) -> str:
    match = re.search(rf'{key}:\s*"([^"]*)"', block)
    return match.group(1) if match else ""


def compact(block: str) -> str:
    return re.sub(r"\s+", " ", block).strip()


def const_block(text: str, name: str, end_marker: str | None = None) -> str:
    marker = f"export const {name}"
    if marker in text:
        start = text.index(marker)
    else:
        start = text.index(f"const {name}")
    brace = text.index("{", start)
    if end_marker and end_marker in text[brace:]:
        return text[brace + 1 : text.index(end_marker, brace)]

    depth = 0
    in_string = False
    escaped = False
    for idx in range(brace, len(text)):
        ch = text[idx]
        if in_string:
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return text[brace + 1 : idx]
    return ""


def top_records(block: str) -> dict[str, str]:
    records: dict[str, str] = {}
    idx = 0
    while idx < len(block):
        key_match = re.search(r'(?:"([^"]+)"|([A-Za-z_]\w*)):\s*\{', block[idx:])
        if not key_match:
            break
        key = key_match.group(1) or key_match.group(2)
        start = idx + key_match.end() - 1
        depth = 0
        in_string = False
        escaped = False
        end = start
        for pos in range(start, len(block)):
            ch = block[pos]
            if in_string:
                if escaped:
                    escaped = False
                elif ch == "\\":
                    escaped = True
                elif ch == '"':
                    in_string = False
                continue
            if ch == '"':
                in_string = True
                continue
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    end = pos
                    break
        records[key] = block[start + 1 : end]
        idx = end + 1
    return records


def clean_text(text: str) -> str:
    replacements = {
        "转机": "可查的缺口",
        "根基": "手里的账路",
        "人心": "旁人的态度",
        "风暴": "急局",
        "成长": "本事",
        "光华": "好感",
        "频率": "动静",
        "曲线": "来路",
        "默认结局": "既定说法",
        "草稿": "旧稿",
        "强力翻盘": "重新开局",
        "大幅提升": "提高",
        "你不再天真，也不浪费机会。": "你要求旧友押下客户名册、担保银，并由商会作见证。",
        "命非天定，亦非人全胜，惟选择可证此生。": "把20关选择刻成名目，留给后来人翻看。"
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = wb.create_sheet(name[:31])
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="3A2815")
    header_font = Font(color="FFF7DF", bold=True)
    thin = Side(style="thin", color="80683A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for idx, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 42))
        ws.column_dimensions[get_column_letter(idx)].width = max(10, min(max_len + 4, 48))

    ws.freeze_panes = "A2"


def export() -> Path:
    life = read("lib/lifeScenarios.ts")
    traits_text = read("lib/destinySkill.ts")
    notes_text = read("lib/choiceOutcomeNotes.ts")
    hidden_text = read("lib/hiddenChoices.ts")
    optimized_text = read("lib/copyOptimizations.ts")

    option_records = top_records(
        const_block(optimized_text, "optimizedOptionCopy", "export const optimizedHiddenStoryTexts")
    )
    hidden_records = top_records(
        const_block(optimized_text, "optimizedHiddenStoryTexts", "export const optimizedTraitCopy")
    )
    hidden_choice_records = top_records(const_block(hidden_text, "hiddenChoiceCopy", "const hiddenStoryTexts"))
    trait_records = top_records(const_block(optimized_text, "optimizedTraitCopy"))

    wb = Workbook()
    wb.remove(wb.active)

    scene_rows: list[list[object]] = []
    choice_rows: list[list[object]] = []
    scenario_blocks = re.findall(
        r'\{\s*\n\s*id:\s*"([^"]+)"([\s\S]*?)(?=\n\s*\},\n\s*\{|\n\s*\}\n\];)',
        life,
    )

    for sid, body in scenario_blocks:
        order = re.search(r"order:\s*(\d+)", body)
        title = qval(body, "title")
        era = qval(body, "era")
        desc_match = re.search(r'description:\s*\n\s*"([\s\S]*?)",\n\s*choices:', body)
        desc = desc_match.group(1) if desc_match else qval(body, "description")
        scene_rows.append([order.group(1) if order else "", sid, title, era, clean_text(desc)])

        choices_match = re.search(r"choices:\s*\[([\s\S]*)\]\s*$", body.strip())
        if choices_match:
            for choice_match in re.finditer(
                r'\{\s*id:\s*"([^"]+)"([\s\S]*?)\}',
                choices_match.group(1),
            ):
                cid = choice_match.group(1)
                cb = choice_match.group(2)
                requirement = re.search(r"requirement:\s*\{([\s\S]*?)\}", cb)
                optimized = option_records.get(f"{sid}:{cid}", {})
                original_label = qval(cb, "label")
                original_desc = qval(cb, "description")
                choice_rows.append(
                    [
                        order.group(1) if order else "",
                        sid,
                        title,
                        cid,
                        original_label,
                        original_desc,
                        qval(cb, "risk") or qval(cb, "riskLevel"),
                        compact(requirement.group(1)) if requirement else "",
                        qval(optimized, "problem"),
                        qval(optimized, "approach"),
                        qval(optimized, "label") or clean_text(original_label),
                        qval(optimized, "description") or clean_text(original_desc),
                        qval(optimized, "motive"),
                        qval(optimized, "cost"),
                        qval(optimized, "benefit"),
                        qval(optimized, "risk"),
                        qval(optimized, "visibleEffect"),
                    ]
                )

    add_sheet(wb, "人生场景", ["序号", "场景ID", "场景名", "年龄阶段", "场景描述"], scene_rows)
    add_sheet(
        wb,
        "普通选项",
        [
            "序号",
            "场景ID",
            "场景名",
            "选项ID",
            "原选项名",
            "原选项描述",
            "风险",
            "解锁条件",
            "原文案问题",
            "优化思路",
            "优化后选项名",
            "优化后选项描述",
            "选择动机",
            "选择代价",
            "可能收益",
            "可能风险",
            "玩家可见效果说明",
        ],
        choice_rows,
    )

    note_rows: list[list[object]] = []
    for match in re.finditer(
        r'"?([\w:-]+)"?:\s*\{\s*\n\s*success:\s*"([^"]*)",\s*\n\s*failure:\s*"([^"]*)"',
        notes_text,
    ):
        key, success, failure = match.groups()
        scenario_id = ""
        choice_id = key
        if ":" in key:
            scenario_id, choice_id = key.split(":", 1)
        note_rows.append(
            [
                scenario_id,
                choice_id,
                success,
                failure,
                "检查是否只有情绪或概念；去掉空泛词后保留具体事件。",
                "保留谁做了什么、损失什么、谁的态度改变。",
                clean_text(success),
                clean_text(failure),
                "成功收益见优化成功后续；失败损失见优化失败后续。",
            ]
        )
    add_sheet(
        wb,
        "普通选项后续",
        [
            "场景ID",
            "选项ID",
            "原成功后续",
            "原失败后续",
            "原文案问题",
            "优化思路",
            "优化版成功后续",
            "优化版失败后续",
            "玩家可见效果说明",
        ],
        note_rows,
    )

    hidden_rows: list[list[object]] = []
    for sid, block in hidden_records.items():
        choice_copy = hidden_choice_records.get(sid, "")
        hidden_rows.append(
            [
                sid,
                qval(choice_copy, "label"),
                qval(choice_copy, "description"),
                qval(block, "greatSuccess"),
                qval(block, "success"),
                qval(block, "partialFail"),
                qval(block, "fail"),
                qval(block, "criticalFail"),
                "已按5档补齐，强调额外资源、正常解决、半成隐患、明确损失、额外后果。",
            ]
        )
    add_sheet(
        wb,
        "隐藏选项后续",
        ["场景ID", "隐藏选项名", "隐藏选项描述", "大成功", "成功", "不尽人意", "失败", "大失败", "优化说明"],
        hidden_rows,
    )

    trait_rows: list[list[object]] = []
    for match in re.finditer(r'\{\s*\n\s*id:\s*"([^"]+)"([\s\S]*?)\n\s*\}', traits_text):
        tid = match.group(1)
        block = match.group(2)
        name = qval(block, "name")
        if not name:
            continue
        stat_bonus = re.search(r"statBonus:\s*\{([\s\S]*?)\}", block)
        tags = re.search(r"tags:\s*\[([\s\S]*?)\]", block)
        optimized = trait_records.get(tid, "")
        trait_rows.append(
            [
                tid,
                name,
                qval(block, "rarity"),
                qval(block, "polarity"),
                qval(block, "description"),
                qval(block, "impact"),
                qval(optimized, "problem"),
                qval(optimized, "approach"),
                qval(optimized, "description") or clean_text(qval(block, "description")),
                qval(optimized, "impact") or clean_text(qval(block, "impact")),
                compact(stat_bonus.group(1)) if stat_bonus else "",
                compact(tags.group(1)) if tags else "",
            ]
        )
    add_sheet(
        wb,
        "命格词条",
        [
            "词条ID",
            "词条名",
            "稀有度",
            "倾向",
            "原短描述",
            "原影响文案",
            "原文案问题",
            "优化思路",
            "优化版短描述",
            "优化版影响文案",
            "属性加成",
            "标签",
        ],
        trait_rows,
    )

    add_sheet(
        wb,
        "说明",
        ["项目", "内容"],
        [
            ["导出时间", "2026-06-10"],
            ["用途", "用于继续交给 GPT 或人工优化游戏文案，保留原文和优化列方便对比。"],
            ["本轮重点", "去套路化、去空话、增加具体动作、代价、收益和失败损失。"],
            ["游戏接入", "优化选项文案、隐藏选项5档后续、命格词条已接入游戏运行层。"],
        ],
    )

    try:
        wb.save(OUT)
        return OUT
    except PermissionError:
        fallback = ROOT / "游戏文案整理-优化版.xlsx"
        wb.save(fallback)
        return fallback


if __name__ == "__main__":
    print(export())
